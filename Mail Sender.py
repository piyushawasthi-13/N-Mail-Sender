from tkinter import *
import tkinter.messagebox
from openpyxl import load_workbook
import smtplib 
from email.mime.multipart import MIMEMultipart 
from email.mime.text import MIMEText 
from email.mime.base import MIMEBase 
from email import encoders 
win=Tk()
win.title("MAIL SENDER:")
win.geometry("550x600+120+120")
fi=StringVar()
mail=StringVar()
passw=StringVar()
def openfile():
    from tkinter import filedialog
    filename2 = filedialog.askopenfilename()
    if filename2.endswith('xlsx'):
        x = load_workbook(filename2)
        s = x.active
        global a
        a = []
        m_row = s.max_row
        m_col = s.max_column
        for i in range(1, m_row+1):
            for j in range(1, m_col+1):
                cell_obj = s.cell(row = i, column = j)
                v = cell_obj.value
                if v.endswith("@gmail.com"):
                    a.append(v)
        tkinter.messagebox.showinfo('Sucess',"File selected : {}".format(filename2))

def new_window():
    show=emailm.get()
    if emailm.get()=="" or passw.get=="":
        tkinter.messagebox.showwarning('denied! ','please fill your detail for login:')
    else:
        win.destroy()
        global top
        top=Tk()
        top.title("MAIL SENDER:")
        top.geometry("550x600+120+120")
        top.geometry("550x600+120+120")
        title=Label(top,text="COMPOSE",width=45,font=("bold",40),fg="brown",)
        title.place(x=-430,y=35)
        pr=Label(top,text=show,font=20)
        pr.place(x=140,y=106)
        xl=Button(top,text="Choose File",font=24,fg="white",bg="brown",command=openfile)
        xl.place(x=60,y=165)
        
        sub=Label(top,text="Subject ",font=20)
        sub.place(x=60,y=216)
        global sub_e
        sub_e=Entry(top,width=20,font=20)
        sub_e.place(x=235,y=218)
        msg=Label(top,text="Message ",font=20)
        msg.place(x=60,y=266)
        global msg1
        msg1=Text(top,width=20,font=5,height=8,wrap=WORD)
        msg1.place(x=235,y=268)
        pathlabel =Label(top)
        pathlabel.pack()
        submit_btn2=Button(top,text="Send",width=8,fg="white",bg="blue",font=5,command=send)
        submit_btn2.place(x=362,y=500)
        xl=Button(top,text="Attach file",font=24,fg="white",bg="brown",command=attach)
        xl.place(x=60,y=500)

win.title("MAIL SENDER:")
win.geometry("550x600")
title=Label(win,text="LOGIN",width=45,font=("bold",40),fg="brown",)
title.place(x=-400,y=70)
email=Label(win,text="Email Id:",font=20)
email.place(x=100,y=200)
emailm=Entry(win,width=25,font=20,textvariable=mail)
emailm.place(x=200,y=200)
password=Label(win,text="password:",font=20)
password.place(x=87,y=250)
password=Entry(win,width=25,font=20,show='*',textvariable=passw)
password.place(x=200,y=250)
submit_btn=Button(win,text="Log In",width=20,fg="white",bg="brown",font=6,command=new_window)
submit_btn.pack()
submit_btn.place(x=200,y=350)
def openfile():
    from tkinter import filedialog
    filename2 = filedialog.askopenfilename()
    xl_label=Label(top,width=20,font=20,text=filename2.split('/')[-1])
    xl_label.place(x=235,y=168)
    if filename2.endswith('xlsx'):
        x = load_workbook(filename2)
        s = x.active
        global a
        a = []
        m_row = s.max_row
        m_col = s.max_column
        for i in range(1, m_row+1):
            for j in range(1, m_col+1):
                cell_obj = s.cell(row = i, column = j)
                v = cell_obj.value
                if v.endswith("@gmail.com"):
                    a.append(v)
        tkinter.messagebox.showinfo('Sucess',"File selected : {}".format(filename2))
def attach():
    
    
    # instance of MIMEMultipart 
    msg = MIMEMultipart()

    msg['From']  = mail.get() 
    msg['Subject'] = sub_e.get()
    # string to store the body of the mail 
    body = msg1.get('1.0', 'end-1c')
    
    # attach the body with the msg instance 
    msg.attach(MIMEText(body, 'plain')) 
    from tkinter import filedialog
    from openpyxl import load_workbook
    filename2 = filedialog.askopenfilename()
    # oen the file to be sent
    if filename2:
        tkinter.messagebox.showinfo('Sucess')  
    attachment = open(filename2, "rb")   
    # instance of MIMEBase and named as p 
    p = MIMEBase('application', 'octet-stream') 
    
    # To change the payload into encoded form 
    p.set_payload((attachment).read()) 
    
    # encode into base64 
    encoders.encode_base64(p) 

    
    p.add_header('Content-Disposition', "attachment; filename= %s" % filename2) 
    print(filename2)
    msg.attach(p)
    global text
    text= msg.as_string() 
    print(filename2)
    print("success")
def send():
    try:
        import smtplib as s
        ob=s.SMTP("smtp.gmail.com",587)
        ob.starttls()
        ob.login(mail.get(),passw.get())
        # # print(message)
        # l=["piyushawasthi131@gmail.com"]
        ob.sendmail(mail.get(),a,text)
        print("send succesfully:")
        tkinter.messagebox.showinfo('Sucess',"sent sucessful")
        ob.quit()     
    except:
        import smtplib as s
        ob=s.SMTP("smtp.gmail.com",587)
        ob.starttls()
        ob.login(mail.get(),passw.get())
        subject=sub_e.get()
        body=msg1.get('1.0', END)
        message="subject:{}\n\n{}".format(subject,body)
        ob.sendmail(mail.get(),a,message)
        print("send succesfully:")
        tkinter.messagebox.showinfo('Sucess',"sent sucessful")
        
        ob.quit()     
win.mainloop()